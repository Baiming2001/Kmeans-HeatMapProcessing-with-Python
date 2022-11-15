# -*- coding: utf-8 -*-
"""
Created on Mon Apr 25 11:04:25 2022

@author: Zeng baiming
"""

import openpyxl,os
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.image as mpimg

"""
   数据导入，数据存储于‘Jiading.xlsx’
"""
      
def cores_classify(a,b):#用于导入质心数据，a为行数，b为列数,暂时不用
    global x_cores,y_cores,cores
    x_cores,y_cores,cores=[],[],[]
    workbook_path= os.path.join('Jiading_logistics.xlsx')
    wb=openpyxl.load_workbook(workbook_path)
    ws=wb["Sheet1"]
    x1,y1=[],[];
    for i in range(1,a+1):
        x=round(ws.cell(row = i, column = b).value,5)
        x1.append(x)
        y=round(ws.cell(row = i, column = b+1).value,5)
        y1.append(y) 
    wb.save(workbook_path)
    x_cores=np.array(x1)
    y_cores=np.array(y1)
    cores.append(np.stack((x_cores,y_cores), axis=1))
    cores=np.vstack(cores)       

def classify(a,b,c):
    #用于导入data数据，排除人口为0的方格。a为行数，b为列数，c为数据文件名。
    #data以最左下角为起点，其中n行m列的人口为data[a*(m-1)+n-1][2]（删除人口为零的方格前）
    global x_width,y_width,x_min1,y_min1
    x_data,y_data,data,population=[],[],[],[]
    workbook_path= os.path.join(c)
    wb=openpyxl.load_workbook(workbook_path)
    ws=wb["Sheet1"]
    x1,y1,z1=[],[],[];
    for q in range(1,b+1):
        for i in range(1,a+1):
            x=round(ws.cell(row = a+1-i, column = q).value,5)
            y=round(ws.cell(row = 2*a+2-i, column = q).value,5)
            z=round(ws.cell(row = a+1-i, column = q+b+1).value)
            if z!=0:
                z1.append(z)  
                y1.append(y) 
                x1.append(x)                  
    wb.save(workbook_path)  
    x_data=np.array(x1)
    y_data=np.array(y1)
    population=np.array(z1)
    x_width,y_width=np.ptp(x_data),np.ptp(y_data)
    x_min1,y_min1=np.min(x_data),np.min(y_data)
    x_data=np.around((x_data-np.min(x_data))/np.ptp(x_data),4)
    y_data=np.around((y_data-np.min(y_data))/np.ptp(y_data),4)
    population=np.around(((population-np.min(population))/np.ptp(population)),4)
    data.append(np.stack((x_data,y_data,population), axis=1))
    data=np.vstack(data)
    return data

def get_point(a,x_min,x_max,y_min,y_max,s):#a为行数，s为文件名
    workbook_path= os.path.join(s)
    wb=openpyxl.load_workbook(workbook_path)
    ws=wb["Sheet1"]
    x1,y1=[],[]
    for i in range(1,a+1):
        x=ws.cell(row = i, column = 9).value#纬度
        y=ws.cell(row = i, column = 10).value#经度
        x1.append(round((x-x_min)/(x_max-x_min)*n_x,0))
        y1.append(round((1-(y-y_min)/(y_max-y_min))*n_y,0))
    return x1,y1
    
def pixel(s,x_min,x_max,y_min,y_max):
    length=s.shape[0]
    for i in range(length):
        s[i][0]=round((s[i][0]-x_min)/(x_max-x_min)*n_x,0)
        s[i][1]=round((1-(s[i][1]-y_min)/(y_max-y_min))*n_y,0)
"""
    k-means聚类算法    
    k       - 指定分簇数量
    ds      - ndarray(m, n)，m个样本的数据集，每个样本n个属性值
"""
def kmeans_xufive(ds, k):  
    m, n = ds.shape # m：样本数量，n：每个样本的属性值个数
    result = np.empty(m, dtype=np.int) # m个样本的聚类结果
    cores = ds[np.random.choice(np.arange(m), k, replace=False)] # 从m个数据样本中不重复地随机选择k个样本作为质心
    average,cost=0,0
    d = np.square(np.repeat(ds, k, axis=0).reshape(m, k, n) - cores)
    distance = np.sqrt(np.sum(d, axis=2)) # ndarray(m, k)，每个样本距离k个质心的距离，共有m行
    index_min = np.argmin(distance, axis=1) # 每个样本距离最近的质心索引序号
        
    while True: # 迭代计算
        d = np.square(np.repeat(ds, k, axis=0).reshape(m, k, n) - cores)
        distance = np.sqrt(np.sum(d, axis=2)) # ndarray(m, k)，每个样本距离k个质心的距离，共有m行
        index_min = np.argmin(distance, axis=1) # 每个样本距离最近的质心索引序号
        if (index_min == result).all(): # 如果样本聚类没有改变
            for i in range(m):
                num=index_min[i]
                cost+=distance[i][num]
            average=cost/m
            for i in range(k):
                cores[i][0]=cores[i][0]*x_width+x_min1
                cores[i][1]=cores[i][1]*y_width+y_min1
            return result,cores,average # 则返回聚类结果和质心数据
        
        result[:] = index_min # 重新分类
        for i in range(k): # 遍历质心集
            items = ds[result==i] # 找出对应当前质心的子样本集
            cores[i] = np.mean(items, axis=0) # 以子样本集的均值作为当前质心的位置

"""
   运行和绘图
"""   
ave,num_up=[],20#设置聚类数num_up
#imread 得到每个像素组成的三维array：（纬度：从上到下，经度：从左到右，维度）
#########################################
#不用重复读取
pic = mpimg.imread('嘉定.png')
n = pic.shape
#########################################
for k in range(10,11):
    ################################
    #设定图片属性
    heighV = n[0]/1000
    widthV = n[1]/1000
    plt.figure(figsize=(widthV, heighV))
    plt.gca().xaxis.set_major_locator(plt.NullLocator())
    plt.gca().yaxis.set_major_locator(plt.NullLocator())
    plt.subplots_adjust(top=1, bottom=0, left=0, right=1, hspace=0, wspace=0)
    plt.margins(0, 0)
    plt.axis('off')
    #################################
    n_y,n_x=n[0],n[1]
    x,y=get_point(94,121.1068,121.3809,31.2271,31.4968,'Santongyida.xlsx')
    plt.scatter(x,y,10,'b','.')
    data =classify(18,16,'Jiading.xlsx')
    result, result_cores,average = kmeans_xufive(data, k)
    print(result_cores)
    ave.append(average)
    pixel(result_cores,121.1068,121.3809,31.2271,31.4968)
    # color = plt.cm.Set1(k)
    # plt.scatter(data[:,0], data[:,1], s=5, c=(result.astype(np.int)))
    plt.scatter(result_cores[:,0], result_cores[:,1], marker='*', c='r')
    k=str(k)
    imgplot=plt.imshow(pic)
    plt.title('Number of cluster center is '+k)
    plt.savefig(k+'-.png',dpi=300)
    #################################
    #如果要保存图片，不用每次都显示
    # plt.show()
    #################################
x = range(2,num_up+1)
plt.figure()
plt.bar(x, ave, width=0.4, label='Bar_x_y' ) #这里是bar()函数
plt.legend()
plt.grid(True, linestyle='--', alpha=0.5)
plt.title('Average cost')
plt.savefig('Average cost'+'.png', dpi=300)
#################################
# plt.show()
#################################
"""
需要自定义的内容有：pixel,pic,get_point,num_up
"""
